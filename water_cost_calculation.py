import os

import openpyxl


def calculate_water_usage(work_book, work_sheet_names, row_num):
    """
    2か月ごとの水道使用量を取得
    """
    # 初期処理
    work_sheet = work_book.get_sheet_by_name(work_sheet_names[0])
    # 水道使用量リスト
    sum_of_water_usage_per_one_month = []
    water_usage_per_one_month = [0]
    """
    index 0 : 3 月
          1 : 4 月
          2 : 5 月
          3 : 6 月
          .
          11 : 2 月
          12 : 3 月
    """
    water_usage_per_two_month = []
    """
    index 0 : 4~5  月
          1 : 6~7　月
          2 : 8~9　月
          3 : 10~11 月
          4 : 12~1　月
          5 : 2~3　月
    """
    # 各月までの累計水道使用量を取得
    for i in range(13):
        cell_num = work_sheet.cell(row=row_num,column=i+3).value
        sum_of_water_usage_per_one_month.append(cell_num)
    # 各月の累計使用量を取得
    for i in range(1,13):
        result = (sum_of_water_usage_per_one_month[i] - sum_of_water_usage_per_one_month[i-1])
        water_usage_per_one_month.append(result)
    # 2か月ごとの累計使用量を取得
    for i in range(1,13,2):
        result = (water_usage_per_one_month[i] + water_usage_per_one_month[i+1])
        water_usage_per_two_month.append(result)
    # 戻り値
    return water_usage_per_two_month

def calculate_sum_water_fee(work_book, work_sheet_names, water_usage_list):
    """
    2か月ごとの水道料金を取得
    """
    # 初期処理
    work_sheet = work_book.get_sheet_by_name(work_sheet_names[1])
    water_fee_list = []
    # 水道料金マスタから水道料金を取得
    for i in range(6):
        for j in range(6,236):
            water_usage = work_sheet.cell(row=j,column=1).value
            if water_usage == water_usage_list[i]:
                water_fee_list.append(work_sheet.cell(row=j,column=4).value)
                break
    return water_fee_list

def water_charge_statement_create(work_book,
                                  work_sheet_names,
                                  row_num,
                                  water_usage_list,
                                  water_fee_list):
    """
    帳票作成
    """
    # 初期処理
    work_sheet1 = work_book.get_sheet_by_name(work_sheet_names[0])
    work_sheet3 = work_book.get_sheet_by_name(work_sheet_names[2])
    # 部屋番号取得
    room_num = work_sheet1.cell(row=row_num,column=1).value
    # 氏名取得
    room_owner = work_sheet1.cell(row=row_num,column=2).value
    # 帳票作成
    work_sheet3.cell(row=5,column=2).value = str(room_num) + '号'
    work_sheet3.cell(row=5,column=3).value = str(room_owner) + '様' 
    for i in range(6):
        work_sheet3.cell(row=10+i,column=4).value = water_usage_list[i]
        work_sheet3.cell(row=10+i,column=5).value = water_fee_list[i]
    # 名前を付けて保存
    work_book.save(f'{work_sheet_names[0]}/{room_owner}.xlsx')

def report_print():
    """
    帳票印刷
    """
    pass
    
#----------------------------------#
#          main 処理
#----------------------------------#
def main():
    book_name = 'water_fee_table.xlsx'
    try:
        # ワークブック取得
        work_book = openpyxl.load_workbook(book_name)
        # ワークシート取得
        work_sheet_names = work_book.get_sheet_names()
        """
        work_sheet[0] : H__年度
        work_sheet[1] : 水道料金マスタ
        work_sheet[2] : 印刷テンプレート
        """
        # 今年度帳票保存ディレクトリ作成
        os.mkdir(work_sheet_names[0])
        # 初期処理
        work_sheet = work_book.get_sheet_by_name(work_sheet_names[0])
        max_row = work_sheet.max_row - 1
        # 一人ずつ帳票作成
        for row_num in range(2,max_row+2):
            # 水道使用量計算
            water_usage_list = calculate_water_usage(work_book, work_sheet_names, row_num)
            # 水道使用料計算
            water_fee_list = calculate_sum_water_fee(work_book, work_sheet_names, water_usage_list)
            # 水道使用量・水道料金明細ファイル作成
            water_charge_statement_create(work_book,
                                          work_sheet_names,
                                          row_num,
                                          water_usage_list,
                                          water_fee_list)
            # 帳票印刷
            report_print()
    except:
        assert False, '処理が中止されました。'
    finally:
        work_book.close


if __name__ == '__main__':
    main()
